"""
Plataforma de Bienestar Psicologico - Unidad de Igualdad de Genero
Universidad Politecnica de Chiapas
"""

import os
import re
import json
import secrets
import openpyxl
from datetime import datetime, timedelta
from functools import wraps
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from flask import (
    Flask, render_template, redirect, url_for, request, send_from_directory,
    flash, session, jsonify, send_file, abort
)

from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager, UserMixin, login_user, logout_user,
    login_required, current_user
) # <-- Aquí faltaba cerrar el paréntesis

from werkzeug.security import generate_password_hash, check_password_hash
# 1. PRIMERO: Crear la aplicación Flask
app = Flask(__name__)

with app.app_context():
    db.create_all()
    print("Tablas creadas exitosamente")

# 2. SEGUNDO: Configurar la aplicación
app.config['SECRET_KEY'] = secrets.token_hex(32)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///bienestar.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=2)

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Por favor inicia sesión para acceder.'
login_manager.login_message_category = 'info'

# --- MODELOS ---

class Usuario(UserMixin, db.Model):
    __tablename__ = 'usuarios'
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    rol = db.Column(db.String(20), nullable=False, default='alumno')
    nombre_completo = db.Column(db.String(200))
    edad = db.Column(db.Integer)
    carrera = db.Column(db.String(100))
    discapacidad = db.Column(db.String(200))
    sexo = db.Column(db.String(20))
    origen = db.Column(db.String(100))
    lengua = db.Column(db.String(100))
    acepta_privacidad = db.Column(db.Boolean, default=False)
    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)
    notificado_riesgo = db.Column(db.Boolean, default=False)
    
    # Relaciones
    resultados_ryff = db.relationship('ResultadoRyff', backref='usuario', lazy=True)
    entradas_diario = db.relationship('EntradaDiario', backref='usuario', lazy=True)
    logros = db.relationship('Logro', backref='usuario', lazy=True)
    notificaciones = db.relationship('Notificacion', backref='usuario', lazy=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class ResultadoRyff(db.Model):
    __tablename__ = 'resultados_ryff'
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    autoaceptacion = db.Column(db.Integer, default=0)
    relaciones_positivas = db.Column(db.Integer, default=0)
    autonomia = db.Column(db.Integer, default=0)
    dominio_entorno = db.Column(db.Integer, default=0)
    crecimiento_personal = db.Column(db.Integer, default=0)
    proposito_vida = db.Column(db.Integer, default=0)
    puntaje_total = db.Column(db.Integer, default=0)
    nivel_riesgo = db.Column(db.String(20))
    respuestas_json = db.Column(db.Text)


class Notificacion(db.Model):
    __tablename__ = 'notificaciones'
    id = db.Column(db.Integer, primary_key=True)
    # Corregido: Referencia a 'usuarios.id' para coincidir con el __tablename__ de Usuario
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    mensaje = db.Column(db.String(500), nullable=False)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    leida = db.Column(db.Boolean, default=False)


class EntradaDiario(db.Model):
    __tablename__ = 'entradas_diario'
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)
    contenido = db.Column(db.Text, nullable=False)
    estado_animo = db.Column(db.Integer, default=3)


class Logro(db.Model):
    __tablename__ = 'logros'
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    titulo = db.Column(db.String(200), nullable=False)
    tipo = db.Column(db.String(50))
    completado = db.Column(db.Boolean, default=False)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)


class FraseConsejo(db.Model):
    __tablename__ = 'frases_consejos'
    id = db.Column(db.Integer, primary_key=True)
    texto = db.Column(db.Text, nullable=False)
    tipo = db.Column(db.String(20))
    activo = db.Column(db.Boolean, default=True)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)

# LOADER & DECORATORS


@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))


def admin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if current_user.rol != 'admin':
            abort(403)
        return f(*args, **kwargs)
    return decorated


# PREGUNTAS TEST DE RYFF (39 items, 6 categorias)


RYFF_CATEGORIAS = {
    'autoaceptacion': {
        'nombre': 'Autoaceptación',
        'descripcion': 'Evalúa el grado en que te sientes bien contigo mismo/a, aceptando tus cualidades y limitaciones.',
        'preguntas': [
            {'id': 1, 'texto': 'Cuando repaso la historia de mi vida, estoy contento/a con como han resultado las cosas.', 'invertida': False},
            {'id': 2, 'texto':'En general, me siento seguro/a y positivo/a conmigo mismo/a.', 'invertida': False},
            {'id': 3, 'texto': 'Siento que muchas de las personas que conozco han sacado mas provecho de la vida que yo.', 'invertida': True},
            {'id': 4, 'texto': 'Me gusta la mayor parte de los aspectos de mi personalidad.', 'invertida': False},
            {'id': 5, 'texto': 'En muchos aspectos, me siento decepcionado/a de mis logros en la vida.', 'invertida': True},
            {'id': 6, 'texto': 'En su mayor parte, me siento orgulloso/a de quien soy y de la vida que llevo.', 'invertida': False},
        ]
    },
    'relaciones_positivas': {
        'nombre': 'Relaciones Positivas',
        'descripcion': 'Mide la calidad de tus relaciones interpersonales y tu capacidad para mantener vínculos de confianza.',
        'preguntas': [
            {'id': 7, 'texto': 'Siento que mis amistades me aportan muchas cosas.', 'invertida': False},
            {'id': 8, 'texto': 'A menudo me siento solo/a porque tengo pocos amigos intimos con los que compartir mis preocupaciones.', 'invertida': True},
            {'id': 9, 'texto': 'No tengo muchas personas que quieran escucharme cuando necesito hablar.', 'invertida': True},
            {'id': 10, 'texto': 'Me parece que la mayor parte de las personas tienen más amigos que yo.', 'invertida': False},
            {'id': 11, 'texto': 'He experimentado muchas relaciones cercanas y de confianza.', 'invertida': False},
            {'id': 12, 'texto': 'Sé que puedo confiar en mis amigos, y ellos/as saben que pueden confiar en mí..', 'invertida': False},
        ]
    },
    'autonomia': {
        'nombre': 'Autonomia',
        'descripcion': 'Evalúa tu independencia, autodeterminación y capacidad para resistir presiones sociales.',
        'preguntas': [
            {'id': 13, 'texto': 'Expreso fácilmente mis opiniones, incluso cuando son opuestas a las opiniones de la mayoría de las personas.', 'invertida': False},
            {'id': 14, 'texto': 'Tiendo a preocuparme por lo que los demás piensan de mí.', 'invertida': True},
            {'id': 15, 'texto': 'Tiendo a dejarme influenciar por personas con opiniones firmes.', 'invertida': True},
            {'id': 16, 'texto': 'Tengo confianza en mis opiniones incluso si son contrarias al consenso general.', 'invertida': False},
            {'id': 17, 'texto': 'Me resulta difícil expresar mis opiniones en temas controvertidos.', 'invertida': True},
            {'id': 18, 'texto': 'A menudo cambio mis decisiones si mis amigos o mi familia están en desacuerdo.', 'invertida': True}
        ]
    },
    'dominio_entorno': {
        'nombre': 'Dominio del Entorno',
        'descripcion': 'Mide tu capacidad para manejar las actividades diarias y crear contextos favorables para ti.',
        'preguntas': [
            {'id': 19, 'texto': 'Me preocupa cómo otras personas evalúan las elecciones que he hecho en mi vida.', 'invertida': True},
            {'id': 20, 'texto': 'Me juzgo por lo que pienso que es importante, no por lo que otros piensan que es importante.', 'invertida': False},
            {'id': 21, 'texto': 'En general, siento que soy responsable de la situación en la que vivo.', 'invertida': False},
            {'id': 22, 'texto': 'Las demandas de la vida diaria a menudo me deprimen.', 'invertida': True},
            {'id': 23, 'texto': 'Soy bastante eficiente, manejando mis responsabilidades diarias.', 'invertida': False},
            {'id': 24, 'texto': 'Si no fuera feliz en mi vida, tomaría medidas efectivas para cambiarla.', 'invertida': False}
        ]
    },
    'crecimiento_personal': {
        'nombre': 'Crecimiento Personal',
        'descripcion': 'Evalúa tu sensacion de desarrollo continuo, apertura a nuevas experiencias y crecimiento como persona.',
        'preguntas': [
            {'id': 25, 'texto': 'Soy una persona activa cuando realizo los proyectos que me propongo.', 'invertida': False},
            {'id': 26, 'texto': 'Conforme pasa el tiempo siento que sigo aprendiendo más sobre mí mismo.', 'invertida': False},
            {'id': 27, 'texto': 'Hace mucho tiempo; que dejé de hacer cambios importantes en mi vida.', 'invertida': True},
            {'id': 28, 'texto': 'Siento que con el tiempo me he desarrollado mucho como persona.', 'invertida': False},
            {'id': 29, 'texto': 'No quiero intentar nuevas formas de hacer las cosas; mi vida está bien como está.', 'invertida': True},
            {'id': 30, 'texto': 'Las experiencias nuevas me desafían a replantear lo que pienso sobre mí mismo y el mundo.', 'invertida': False},
            {'id': 31, 'texto': 'Pensándolo bien, con los años no he mejorado mucho como persona.', 'invertida': True},
            {'id': 32, 'texto': 'Para mí, la vida ha sido un proceso continuo de aprendizaje, desarrollo y crecimiento.', 'invertida': False}
        ]
        },
    'proposito_vida': {
        'nombre': 'Propósito en la Vida',
        'descripcion': 'Mide el sentido de dirección, metas claras y la sensación de que la vida tiene significado.',
        'preguntas': [
            {'id': 33, 'texto': 'Me resulta difícil dirigir mi vida hacia un camino que me satisfaga.', 'invertida': True},
            {'id': 34, 'texto': 'Disfruto haciendo planes para el futuro y trabajar para hacerlos realidad.', 'invertida': False},
            {'id': 35, 'texto': 'Soy parte de una familia y he construido un modo (estilo) de vida a mi gusto.', 'invertida': False},
            {'id': 36, 'texto': 'Me siento bien cuando pienso lo que he hecho en el pasado y lo que espero hacer en el futuro.', 'invertida': False},
            {'id': 37, 'texto': 'Mis objetivos en la vida han sido más una fuente de satisfacción que de frustración para mí.', 'invertida': False},
            {'id': 38, 'texto': 'Tengo clara la dirección y objetivo de mi vida.', 'invertida': False},
            {'id': 39, 'texto': 'No tengo claro qué es lo que intento conseguir en la vida.', 'invertida': True}
        ]
    },
}

RYFF_INTERPRETACIONES = {
    'alto': {
        'autoaceptacion': 'Tienes una actitud positiva hacia ti mismo/a. Reconoces y aceptas tus múltiples aspectos, incluyendo las cualidades buenas y malas. Te sientes bien acerca de tu pasado.',
        'relaciones_positivas': 'Mantienes relaciones calidas, satisfactorias y de confianza con los demas. Te preocupas por el bienestar de otros y muestras empatía, afecto e intimidad.',
        'autonomia': 'Eres autodeterminado/a e independiente. Resistes las presiones sociales y regulas tu comportamiento desde dentro. Te evalúas con criterios personales.',
        'dominio_entorno': 'Tienes un sentido de dominio y competencia manejando tu entorno. Usas de forma eficaz las oportunidades que te rodean y eres capaz de elegir contextos adecuados.',
        'crecimiento_personal': 'Tienes un sentimiento de desarrollo continuo. Te ves creciendo y expandiendote. Estas abierto/a a nuevas experiencias.',
        'proposito_vida': 'Tienes metas claras y un sentido de dirección en la vida. Sientes que tu vida presente y pasada tienen significado. Tienes creencias que te dan un propósito.'
    },
    'medio': {
        'autoaceptacion': 'Tienes una aceptación moderada de ti mismo/a. En general te sientes bien, pero hay aspectos que te gustaría mejorar o que a veces te generan insatisfacción.',
        'relaciones_positivas': 'Tus relaciones sociales son estables pero podrian profundizarse. Tienes conexiones pero a veces puedes sentir que falta mayor cercanía o confianza.',
        'autonomia': 'Tienes un nivel moderado de independencia. A veces te dejas influenciar por las opiniones de otros pero en general mantienes tus criterios.',
        'dominio_entorno': 'Manejas las situaciones de la vida cotidiana de forma aceptable, aunque a veces sientes que podrías organizar mejor tu entorno o aprovechar mas las oportunidades.',
        'crecimiento_personal': 'Tienes cierto sentido de crecimiento pero podrías buscar mas activamente nuevas experiencias y desafíos para tu desarrollo personal.',
        'proposito_vida': 'Tienes algunas metas y sentido de dirección, pero a veces puedes sentir incertidumbre sobre tu propósito o hacia donde te diriges.'
    },
    'bajo': {
        'autoaceptacion': 'Podrías estar experimentando insatisfacción contigo mismo/a. Es posible que desees ser diferente o que te perturben ciertas cualidades personales. Te recomendamos buscar apoyo.',
        'relaciones_positivas': 'Podrías estar experimentando aislamiento o dificultades en tus relaciones. Es importante buscar conexiones significativas y considerar apoyo profesional.',
        'autonomia': 'Podrías estar dejandote influenciar demasiado por las expectativas y evaluaciones de otros. Trabajar en tu autoconfianza podría beneficiarte.',
        'dominio_entorno': 'Podrías estar sintiendo que la vida diaria te supera. Organizar mejor tus actividades y buscar apoyo podria ayudarte a sentir mayor control.',
        'crecimiento_personal': 'Podrías sentir estancamiento o falta de desarrollo. Buscar nuevas experiencias y objetivos podría ayudarte a crecer como persona.',
        'proposito_vida': 'Podrías estar sintiendo falta de dirección o propósito. Definir metas claras y buscar orientación profesional podria darte mayor sentido.'
    }
}

# RUTAS AUTENTICACION

@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.rol == 'admin':
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('alumno_dashboard'))
    return render_template('auth/login.html')

@app.route('/service-worker.js')
def sw():
    return send_from_directory('static', 'service-worker.js')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        if current_user.rol == 'admin':
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('alumno_dashboard'))

    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        usuario = Usuario.query.filter_by(email=email).first()
        if usuario and usuario.check_password(password):
            login_user(usuario, remember=False)
            session.permanent = True
            if usuario.rol == 'admin':
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('alumno_dashboard'))
        flash('Correo o contraseña incorrectos.', 'error')
    return render_template('auth/login.html')


@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')

        pattern = r'^[a-zA-Z0-9]+@[a-zA-Z]+\.upchiapas\.edu\.mx$'
        if not re.match(pattern, email):
            flash('El correo debe ser institucional (matricula@upchiapas.edu.mx).', 'error')
            return render_template('auth/registro.html')
        if password != confirm_password:
            flash('Las contraseñas no coinciden.', 'error')
            return render_template('auth/registro.html')
        if len(password) < 6:
            flash('La contraseña debe tener al menos 6 caracteres.', 'error')
            return render_template('auth/registro.html')
        if Usuario.query.filter_by(email=email).first():
            flash('Este correo ya esta registrado.', 'error')
            return render_template('auth/registro.html')

        acepta = request.form.get('acepta_privacidad') == 'on'
        if not acepta:
            flash('Debes aceptar el aviso de privacidad para continuar.', 'error')
            return render_template('auth/registro.html')

        nuevo = Usuario(
            email=email, rol='alumno',
            nombre_completo=request.form.get('nombre_completo', '').strip(),
            edad=int(request.form.get('edad', 0)) if request.form.get('edad') else None,
            carrera=request.form.get('carrera', '').strip(),
            discapacidad=request.form.get('discapacidad', '').strip(),
            sexo=request.form.get('sexo', ''),
            origen=request.form.get('origen', '').strip(),
            lengua=request.form.get('lengua', '').strip(),
            acepta_privacidad=acepta
        )
        nuevo.set_password(password)
        db.session.add(nuevo)
        db.session.commit()
        flash('Registro exitoso! Ahora puedes iniciar sesión.', 'success')
        return redirect(url_for('login'))
    return render_template('auth/registro.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    session.clear()
    response = redirect(url_for('login'))
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


# RUTAS ALUMNO

@app.route('/alumno')
@login_required
def alumno_dashboard():
    # 1. Seguridad: Solo alumnos entran aquí
    if current_user.rol != 'alumno':
        abort(403)
    
    # 2. Lógica de Notificaciones (Lo nuevo)
    notificaciones = Notificacion.query.filter_by(usuario_id=current_user.id).order_by(Notificacion.fecha.desc()).all()
    
    # 3. Lógica de Estadísticas (Lo que ya tenías)
    ultimo_test = ResultadoRyff.query.filter_by(usuario_id=current_user.id).order_by(ResultadoRyff.fecha.desc()).first()
    total_diario = EntradaDiario.query.filter_by(usuario_id=current_user.id).count()
    logros_completados = Logro.query.filter_by(usuario_id=current_user.id, completado=True).count()
    
    # 4. Enviamos TODO al HTML en un solo paquete
    return render_template('alumno/dashboard.html', 
                        notificaciones=notificaciones, 
                        ultimo_test=ultimo_test, 
                        total_diario=total_diario, 
                        logros_completados=logros_completados)

@app.route('/alumno/test-ryff', methods=['GET'])
@login_required
def test_ryff():
    if current_user.rol != 'alumno':
        abort(403)
    return render_template('alumno/test_ryff.html', categorias=RYFF_CATEGORIAS)


@app.route('/alumno/test-ryff/enviar', methods=['POST'])
@login_required
def enviar_test_ryff():
    if current_user.rol != 'alumno':
        abort(403)
    respuestas = {}
    puntajes = {}
    for cat_key, cat_data in RYFF_CATEGORIAS.items():
        total_cat = 0
        for preg in cat_data['preguntas']:
            val = request.form.get(f'pregunta_{preg["id"]}')
            if val is None:
                flash('Debes responder todas las preguntas.', 'error')
                return redirect(url_for('test_ryff'))
            val = int(val)
            respuestas[str(preg['id'])] = val

            # Ítems inversos: aplicar fórmula (7 - puntaje original)
            if preg['invertida']:
                total_cat += (7 - val)
            else:
                total_cat += val

        puntajes[cat_key] = total_cat

    # Puntaje total: suma de las 6 dimensiones (máximo teórico = 234)
    puntaje_total = sum(puntajes.values())

    # Interpretación por rangos absolutos (NO por porcentaje)
    if puntaje_total >= 170:
        nivel = 'alto'
    elif puntaje_total >= 105:
        nivel = 'medio'
    else:
        nivel = 'bajo'

    resultado = ResultadoRyff(
        usuario_id=current_user.id,
        autoaceptacion=puntajes['autoaceptacion'],
        relaciones_positivas=puntajes['relaciones_positivas'],
        autonomia=puntajes['autonomia'],
        dominio_entorno=puntajes['dominio_entorno'],
        crecimiento_personal=puntajes['crecimiento_personal'],
        proposito_vida=puntajes['proposito_vida'],
        puntaje_total=puntaje_total,
        nivel_riesgo=nivel,
        respuestas_json=json.dumps(respuestas)
    )
    db.session.add(resultado)
    db.session.commit()
    return redirect(url_for('resultado_ryff', resultado_id=resultado.id))

@app.route('/alumno/test-ryff/resultado/<int:resultado_id>')
@login_required
def resultado_ryff(resultado_id):
    if current_user.rol != 'alumno':
        abort(403)
    resultado = ResultadoRyff.query.get_or_404(resultado_id)
    if resultado.usuario_id != current_user.id:
        abort(403)
    categorias_resultados = []
    for cat_key, cat_data in RYFF_CATEGORIAS.items():
        puntaje = getattr(resultado, cat_key)
        max_cat = 42
        porcentaje_cat = (puntaje / max_cat) * 100
        if porcentaje_cat >= 66:
            nivel_cat = 'alto'
        elif porcentaje_cat >= 33:
            nivel_cat = 'medio'
        else:
            nivel_cat = 'bajo'
        categorias_resultados.append({
            'key': cat_key, 'nombre': cat_data['nombre'],
            'descripcion': cat_data['descripcion'],
            'puntaje': puntaje, 'max': max_cat,
            'porcentaje': round(porcentaje_cat, 1),
            'nivel': nivel_cat,
            'interpretacion': RYFF_INTERPRETACIONES[nivel_cat][cat_key]
        })
    return render_template('alumno/resultado_ryff.html', resultado=resultado, categorias=categorias_resultados)


@app.route('/alumno/historial-ryff')
@login_required
def historial_ryff():
    if current_user.rol != 'alumno':
        abort(403)
    resultados = ResultadoRyff.query.filter_by(usuario_id=current_user.id).order_by(ResultadoRyff.fecha.desc()).all()
    return render_template('alumno/historial_ryff.html', resultados=resultados)


@app.route('/alumno/centro-equilibrio')
@login_required
def centro_equilibrio():
    if current_user.rol != 'alumno':
        abort(403)
    return render_template('alumno/centro_equilibrio.html')


@app.route('/alumno/frases')
@login_required
def frases_consejos():
    if current_user.rol != 'alumno':
        abort(403)
    frases = FraseConsejo.query.filter_by(activo=True).all()
    return render_template('alumno/frases.html', frases=frases)


@app.route('/alumno/diario', methods=['GET', 'POST'])
@login_required
def diario():
    if current_user.rol != 'alumno':
        abort(403)
    if request.method == 'POST':
        contenido = request.form.get('contenido', '').strip()
        estado = int(request.form.get('estado_animo', 3))
        if contenido:
            entrada = EntradaDiario(usuario_id=current_user.id, contenido=contenido, estado_animo=estado)
            db.session.add(entrada)
            db.session.commit()
            flash('Entrada guardada.', 'success')
        return redirect(url_for('diario'))
    entradas = EntradaDiario.query.filter_by(usuario_id=current_user.id).order_by(EntradaDiario.fecha.desc()).all()
    datos_grafica = []
    for e in reversed(entradas[-30:]):
        datos_grafica.append({'fecha': e.fecha.strftime('%d/%m'), 'animo': e.estado_animo})
    return render_template('alumno/diario.html', entradas=entradas, datos_grafica=json.dumps(datos_grafica))


@app.route('/alumno/logros', methods=['GET', 'POST'])
@login_required
def logros():
    if current_user.rol != 'alumno':
        abort(403)
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'agregar':
            titulo = request.form.get('titulo', '').strip()
            if titulo:
                db.session.add(Logro(usuario_id=current_user.id, titulo=titulo, tipo='personal'))
                db.session.commit()
        elif action == 'toggle':
            logro = Logro.query.get(request.form.get('logro_id'))
            if logro and logro.usuario_id == current_user.id:
                logro.completado = not logro.completado
                logro.fecha = datetime.utcnow()
                db.session.commit()
        elif action == 'eliminar':
            logro = Logro.query.get(request.form.get('logro_id'))
            if logro and logro.usuario_id == current_user.id:
                db.session.delete(logro)
                db.session.commit()
        return redirect(url_for('logros'))

    mis_logros = Logro.query.filter_by(usuario_id=current_user.id).order_by(Logro.completado, Logro.fecha.desc()).all()
    predefinidos = [
        'Tomar 2 litros de agua', 'Salir a correr', 'Ir al gimnasio',
        'Meditar 10 minutos', 'Leer 30 minutos', 'Dormir 8 horas',
        'Comer frutas y verduras', 'Llamar a un ser querido',
        'Hacer ejercicios de respiración', 'Escribir en mi diario'
    ]
    existentes = [l.titulo for l in mis_logros]
    for p in predefinidos:
        if p not in existentes:
            db.session.add(Logro(usuario_id=current_user.id, titulo=p, tipo='predefinido'))
    db.session.commit()
    mis_logros = Logro.query.filter_by(usuario_id=current_user.id).order_by(Logro.completado, Logro.fecha.desc()).all()
    return render_template('alumno/logros.html', logros=mis_logros)

@app.route('/alumno/cambiar-password', methods=['POST'])
@login_required
def alumno_cambiar_password():
    # Obtenemos los datos del formulario
    cur = request.form.get('current_password', '')
    new = request.form.get('new_password', '')
    conf = request.form.get('confirm_password', '')

    # Validaciones
    if not current_user.check_password(cur):
        flash('La contraseña actual es incorrecta.', 'error')
    elif new != conf:
        flash('Las nuevas contraseñas no coinciden.', 'error')
    elif len(new) < 6:
        flash('La nueva contraseña debe tener al menos 6 caracteres.', 'error')
    else:
        # Si todo está bien, guardamos
        current_user.set_password(new)
        db.session.commit()
        flash('Tu contraseña ha sido actualizada exitosamente.', 'success')
    
    # Redirigimos al dashboard del alumno
    return redirect(url_for('alumno_dashboard'))



# RUTAS ADMIN
@app.route('/admin')
@admin_required
def admin_dashboard():
    # Conteos generales
    total_alumnos = Usuario.query.filter_by(rol='alumno').count()
    total_tests = ResultadoRyff.query.count()
    
    # 1. Conteo para el cuadro de "Riesgo Alto" (Badge rojo)
    # Filtramos por el último test de cada usuario para no contar repetidos
    subquery = db.session.query(
        ResultadoRyff.usuario_id, 
        db.func.max(ResultadoRyff.id).label('max_id')
    ).group_by(ResultadoRyff.usuario_id).subquery()

    alumnos_riesgo_alto = db.session.query(Usuario).join(
        ResultadoRyff, Usuario.id == ResultadoRyff.usuario_id
    ).join(
        subquery, ResultadoRyff.id == subquery.c.max_id
    ).filter(
        ResultadoRyff.nivel_riesgo == 'bajo'
    ).count()

    # Nota: No necesitamos calcular los conteos de la gráfica aquí 
    # porque los pide la ruta /admin/estadisticas-json por separado.

    return render_template('admin/dashboard.html', 
                        total_alumnos=total_alumnos, 
                        total_tests=total_tests, 
                        alumnos_riesgo_alto=alumnos_riesgo_alto)

@app.route('/admin/resultados')
@admin_required
def admin_resultados():
    subquery = db.session.query(
        ResultadoRyff.usuario_id, db.func.max(ResultadoRyff.id).label('max_id')
    ).group_by(ResultadoRyff.usuario_id).subquery()
    
    resultados = db.session.query(ResultadoRyff, Usuario).join(
        subquery, ResultadoRyff.id == subquery.c.max_id
    ).join(Usuario, ResultadoRyff.usuario_id == Usuario.id).order_by(
        # Esto ya lo tienes bien: pone los riesgos ('bajo') hasta arriba
        db.case(
            (ResultadoRyff.nivel_riesgo == 'bajo', 1), 
            (ResultadoRyff.nivel_riesgo == 'medio', 2), 
            else_=3
        )
    ).all()
    return render_template('admin/resultados.html', resultados=resultados)


@app.route('/admin/enviar-alerta/<int:usuario_id>', methods=['POST'])
@admin_required
def enviar_alerta(usuario_id):
    usuario = Usuario.query.get_or_404(usuario_id)
    
    # Creamos la notificación interna en la App
    nueva_notif = Notificacion(
        usuario_id=usuario.id,
        mensaje="Atención: Tu resultado del test indica un nivel de riesgo. Te invitamos a agendar una cita en el SIAUP para seguimiento."
    )
    
    usuario.notificado_riesgo = True
    db.session.add(nueva_notif)
    db.session.commit()
    
    flash(f'Notificación enviada internamente a {usuario.nombre_completo}.', 'success')
    return redirect(url_for('admin_resultados'))


@app.route('/admin/enviar-alertas-masivas', methods=['POST'])
@admin_required
def enviar_alertas_masivas():
    subquery = db.session.query(
        ResultadoRyff.usuario_id, db.func.max(ResultadoRyff.id).label('max_id')
    ).group_by(ResultadoRyff.usuario_id).subquery()
    
    alumnos_en_riesgo = db.session.query(Usuario).join(
        ResultadoRyff, Usuario.id == ResultadoRyff.usuario_id
    ).join(subquery, ResultadoRyff.id == subquery.c.max_id).filter(
        ResultadoRyff.nivel_riesgo == 'bajo', 
        Usuario.notificado_riesgo == False
    ).all()
    
    count = 0

    for u in alumnos_en_riesgo:
        nueva_notif = Notificacion(
            usuario_id=u.id,
            mensaje=(f"Hola {u.nombre_completo}, tras revisar tus resultados del Test de Bienestar, "
                    f"te invitamos a agendar una cita en el SIAUP para brindarte orientación personalizada.")
        )
        db.session.add(nueva_notif)
        u.notificado_riesgo = True
        count += 1
    db.session.commit()
    
    if count > 0:
        flash(f'Éxito: Se enviaron {count} notificaciones internas a alumnos en riesgo alto.', 'success')
    else:
        flash('No se encontraron nuevos alumnos en riesgo para notificar.', 'info')
        
    return redirect(url_for('admin_resultados'))
@app.route('/admin/frases', methods=['GET', 'POST'])
@admin_required
def admin_frases():
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'agregar':
            texto = request.form.get('texto', '').strip()
            tipo = request.form.get('tipo', 'frase')
            if texto:
                db.session.add(FraseConsejo(texto=texto, tipo=tipo))
                db.session.commit()
                flash('Frase/consejo agregado.', 'success')
        elif action == 'eliminar':
            f = FraseConsejo.query.get(request.form.get('frase_id'))
            if f:
                db.session.delete(f)
                db.session.commit()
                flash('Frase/consejo eliminado.', 'success')
        elif action == 'toggle':
            f = FraseConsejo.query.get(request.form.get('frase_id'))
            if f:
                f.activo = not f.activo
                db.session.commit()
        return redirect(url_for('admin_frases'))
    frases = FraseConsejo.query.order_by(FraseConsejo.fecha_creacion.desc()).all()
    return render_template('admin/frases.html', frases=frases)


@app.route('/admin/exportar-excel')
@admin_required
def exportar_excel(): 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados Bienestar"
    
    headers = ['Nombre','Email','Edad','Carrera','Sexo','Origen','Lengua','Discapacidad','Fecha Test','Autoaceptación','Relaciones','Autonomía','Dominio Entorno','Crecimiento','Propósito','Puntaje Total','Nivel Riesgo','Notificado']
    
    hfill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
    hfont = Font(color='FFFFFF', bold=True, size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    subquery = db.session.query(ResultadoRyff.usuario_id, db.func.max(ResultadoRyff.id).label('max_id')).group_by(ResultadoRyff.usuario_id).subquery()
    resultados = db.session.query(ResultadoRyff, Usuario).join(subquery, ResultadoRyff.id == subquery.c.max_id).join(Usuario, ResultadoRyff.usuario_id == Usuario.id).order_by(
        db.case((ResultadoRyff.nivel_riesgo == 'alto', 1), (ResultadoRyff.nivel_riesgo == 'medio', 2), else_=3)).all()

    risk_colors = {
        'alto': PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid'), 
        'medio': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'), 
        'bajo': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    }

    for ri, (r, u) in enumerate(resultados, 2):
        data = [u.nombre_completo, u.email, u.edad, u.carrera, u.sexo, u.origen, u.lengua, u.discapacidad, r.fecha.strftime('%d/%m/%Y %H:%M') if r.fecha else '', r.autoaceptacion, r.relaciones_positivas, r.autonomia, r.dominio_entorno, r.crecimiento_personal, r.proposito_vida, r.puntaje_total, r.nivel_riesgo, 'Si' if u.notificado_riesgo else 'No']
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            if col == 17: # Columna de Nivel de Riesgo
                fill = risk_colors.get(val)
                if fill: cell.fill = fill

    # Ajustar ancho de columnas
    for col in range(1, len(headers)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    # Guardar y enviar archivo
    filepath = os.path.join(os.path.dirname(__file__), 'resultados_bienestar.xlsx')
    wb.save(filepath)
    return send_file(filepath, as_attachment=True, download_name='resultados_bienestar.xlsx')

@app.route('/admin/limpiar-db', methods=['POST'])
@admin_required
def limpiar_db():
    if request.form.get('confirmar') == 'CONFIRMAR':
        ResultadoRyff.query.delete(); EntradaDiario.query.delete(); Logro.query.delete()
        db.session.commit()
        flash('Datos de pruebas eliminados exitosamente.', 'success')
    else:
        flash('Escribe CONFIRMAR para eliminar los datos.', 'error')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/cambiar-password', methods=['POST'])
@admin_required
def admin_cambiar_password():
    cur = request.form.get('current_password', '')
    new = request.form.get('new_password', '')
    conf = request.form.get('confirm_password', '')
    if not current_user.check_password(cur):
        flash('Contraseña actual incorrecta.', 'error')
    elif new != conf:
        flash('Las contrasenñas nuevas no coinciden.', 'error')
    elif len(new) < 6:
        flash('La nueva contraseña debe tener al menos 6 caracteres.', 'error')
    else:
        current_user.set_password(new)
        db.session.commit()
        flash('Contraseña actualizada exitosamente.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/rescate-admin')
def rescate_admin():
    # Buscamos al primer usuario que tenga rol de admin
    admin = Usuario.query.filter_by(rol='admin').first()
    
    if not admin:
        # Si de plano no hay admin, creamos uno nuevo
        admin = Usuario(
            email='igualdaddegenero@upchiapas.edu.mx',
            rol='admin',
            nombre_completo='Administrador Unidad'
        )
        db.session.add(admin)
        mensaje = "Se CREÓ un nuevo administrador desde cero."
    else:
        # Si existe, actualizamos sus datos
        admin.email = 'igualdaddegenero@upchiapas.edu.mx'
        mensaje = "Se ACTUALIZÓ el administrador existente."

    admin.set_password('Admin123') 
    
    db.session.commit()
    return f"{mensaje} <br> Correo: <b>igualdaddegenero@upchiapas.edu.mx</b> <br> Contraseña: <b>Admin123</b> <br><br> Intenta loguearte ahora."


@app.route('/admin/estadisticas-json')
@admin_required
def estadisticas_json():
    # Contamos la realidad de la base de datos
    c_bajo = ResultadoRyff.query.filter_by(nivel_riesgo='bajo').count()
    c_medio = ResultadoRyff.query.filter_by(nivel_riesgo='medio').count()
    c_alto = ResultadoRyff.query.filter_by(nivel_riesgo='alto').count()

    # IMPORTANTE: Mapeamos para que el JS reciba 'alto' y lo pinte ROJO
    return jsonify({
        'riesgo': {
            'alto': c_bajo,    # Bienestar bajo = Se pinta en la sección ROJA
            'medio': c_medio,  # Se pinta en la sección AMARILLA
            'bajo': c_alto     # Bienestar alto = Se pinta en la sección VERDE
        }
    })


# INICIALIZACION

def init_db():
    with app.app_context():
        db.create_all()
        admin = Usuario.query.filter_by(rol='admin').first()
        if not admin:
            admin = Usuario(email='igualdaddegenero@upchiapas.edu.mx', rol='admin', nombre_completo='Administrador - Unidad de Igualdad de Genero')
            admin.set_password('admin2026')
            db.session.add(admin)
        if FraseConsejo.query.count() == 0:
            for texto, tipo in [
                ('Cada día es una nueva oportunidad para crecer y aprender.', 'frase'),
                ('Tu bienestar mental es tan importante como tu bienestar físico.', 'frase'),
                ('No tienes que ser perfecto/a, solo tienes que ser autentico/a.', 'frase'),
                ('Pedir ayuda no es senal de debilidad, es senal de valentia.', 'frase'),
                ('Eres mas fuerte de lo que crees y mas capaz de lo que imaginas.', 'frase'),
                ('Tomate un momento para respirar profundamente y reconectar contigo.', 'consejo'),
                ('Intenta dormir al menos 7 horas cada noche para mejorar tu bienestar.', 'consejo'),
                ('Manten una rutina de ejercicio: 30 minutos diarios hacen la diferencia.', 'consejo'),
                ('Habla con alguien de confianza cuando sientas que lo necesitas.', 'consejo'),
                ('Celebra tus pequenos logros, cada paso cuenta.', 'consejo'),
                ('La igualdad de genero comienza con el respeto mutuo.', 'frase'),
                ('Tu voz importa. No dejes que nadie te haga sentir lo contrario.', 'frase'),
            ]:
                db.session.add(FraseConsejo(texto=texto, tipo=tipo))
        db.session.commit()


if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)